from flask import Flask, render_template, request, redirect, url_for,flash, session, make_response, send_from_directory
import pandas as pd
import requests
import io
from markupsafe import escape

import urllib.parse # Import urllib.parse for URL encoding
from weasyprint import HTML # Import HTML from weasyprint
from werkzeug.security import generate_password_hash, check_password_hash # Import security utilities
from datetime import datetime # Import datetime for PDF timestamp
import os # Import os for file path operations
from werkzeug.utils import secure_filename # Import secure_filename for safe file uploads

# Import libraries for reading document content (install these: pip install pypdf python-docx openpyxl)
try:
    import pypdf # Use pypdf as PyPDF2 is deprecated
except ImportError:
    pypdf = None
    print("Warning: pypdf not installed. PDF content will not be available to AI Assistant.")

try:
    import docx # Use python-docx
except ImportError:
    docx = None
    print("Warning: python-docx not installed. Word document content will not be available to AI Assistant.")

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("Warning: openpyxl not installed. Excel document content will not be available to AI Assistant.")

# Import the google_search tool
# Ensure you have the google_search tool available in your environment
try:
    # from google_search import search
except ImportError:
    search = None
    print("Warning: google_search tool not available. Web search functionality will be limited.")


app = Flask(__name__)
# --- Security Note: In a real application, use a strong, random secret key ---
# Secret key is needed for flashing messages.
app.secret_key = 'your_very_secret_key_replace_this' # Required for flashing messages and sessions

# --- Configuration ---
# Replace with your actual Google Sheet URL
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1JYs4ZtUKfklu-bEqdYOeeKu6nF7rM5I55EQZb-yrs-A/edit?gid=0#gid=0"
# Extract the sheet ID from the URL
SHEET_ID = GOOGLE_SHEET_URL.split('/d/')[1].split('/')[0]
# Construct the export URL for CSV. Assuming the sheet name is 'Sheet1'.
# If your sheet name is different, change 'sheet=Sheet1' accordingly.
CSV_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet=Sheet1"

# Arkesel API Configuration
# IMPORTANT: Double-check that this API key is correct and active in your Arkesel account.
# Using the API key provided by the user for the older endpoint.
ARKESEL_API_KEY = "b0FrYkNNVlZGSmdrendVT3hwUHk"
# Using the older GET-based SMS send URL provided by the user.
ARKESEL_SMS_URL = "https://sms.arkesel.com/sms/api"
# IMPORTANT: Replace with your registered Arkesel Sender ID.
# Verify this Sender ID is registered and approved in your Arkesel account.
ARKESEL_SENDER_ID = "GyeduTech" # e.g., "MySchool"

# --- Website Domain Configuration ---
# IMPORTANT: Replace with your actual website domain (e.g., 'https://your-school-website.com')
WEBSITE_DOMAIN = "https://flask-sms-app.onrender.com/" # Replace with your actual domain in production

# --- Admin Password Hashing ---
# Hash for the password 'gyedu2025'
# In a real application, generate this hash once and store it securely (e.g., in environment variables or a config file).
ADMIN_PASSWORD_HASH = generate_password_hash('gyedu2025') # Hashing the password 'gyedu2025'
print(f"Admin password hash (for 'gyedu2025'): {ADMIN_PASSWORD_HASH}") # Print hash for verification

# --- File Upload Configuration ---
# Create an 'uploads' directory if it doesn't exist
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Document Content Reading Functions ---
# These functions require external libraries (pypdf, python-docx, openpyxl)
# and provide basic text extraction. Full LLM integration with documents
# is complex and requires a different architecture (e.g., RAG pipeline).
# This implementation simulates using the documents by listing them or
# including a placeholder message in the AI response.

def read_pdf(filepath):
    if not pypdf:
        return "PDF reading library not available."
    try:
        with open(filepath, 'rb') as f:
            reader = pypdf.PdfReader(f)
            text = ''
            for page_num in range(len(reader.pages)):
                text += reader.pages[page_num].extract_text() + '\n'
            return text
    except Exception as e:
        print(f"Error reading PDF {filepath}: {e}")
        return f"Error reading PDF file: {e}"

def read_docx(filepath):
    if not docx:
        return "Word document reading library not available."
    try:
        doc = docx.Document(filepath)
        text = ''
        for paragraph in doc.paragraphs:
            text += paragraph.text + '\n'
        return text
    except Exception as e:
        print(f"Error reading DOCX {filepath}: {e}")
        return f"Error reading Word document file: {e}"

def read_excel(filepath):
    if not openpyxl:
        return "Excel reading library not available."
    try:
        workbook = openpyxl.load_workbook(filepath)
        text = ''
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else '' for cell in row]
                text += '\t'.join(row_values) + '\n'
            text += '\n'
        return text
    except Exception as e:
        print(f"Error reading Excel {filepath}: {e}")
        return f"Error reading Excel file: {e}"

def get_document_content(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return None # File not found

    file_extension = filename.rsplit('.', 1)[1].lower()

    if file_extension == 'pdf':
        return read_pdf(filepath)
    elif file_extension == 'docx':
        return read_docx(filepath)
    elif file_extension == 'doc': # Basic handling for .doc (might not work for all formats without additional libraries)
         return "Reading .doc files is not fully supported without additional libraries."
    elif file_extension in ['xls', 'xlsx']:
        return read_excel(filepath)
    else:
        return "Unsupported file type for reading."


# Define the column mapping from your sheet.
# The keys are internal identifiers, the values MUST EXACTLY MATCH the column headers in your Google Sheet.
# This will be populated based on the nested SUBJECT_DETAILS structure.
COLUMN_MAPPING = {
    'Student ID': 'student_id',
    'Student Department': 'department',      # Example: Column header in sheet is 'Student ID'
    'Student Name': 'studentName',  # Example: Column header in sheet is 'Student Name'
    'Parent Phone': 'parentPhone',  # Example: Column header in sheet is 'Parent Phone'
    'Math Score': 'mathematics',      # Example: Column header in sheet is 'Math Score'
    'Science Score': 'science',# Example: Column header in sheet is 'Science Score'
    'Social Score': 'social',
    'Entrepreneur Score': 'entrepreneur',
    'Practicals': 'practicals',
    'Semester': 'semester',
    'Class': 'form',
    'Total Score': 'totalScore',  # Example: Column header in sheet is 'Parent Phone'
    # Subject specific columns will be added here based on SUBJECT_DETAILS
}

# Define the structure for each subject's details PER SEMESTER and their corresponding column headers in the Google Sheet.
# The outer keys are the Semester Names you want to display.
# The inner keys are the Subject Names you want to display.
# The innermost values are dictionaries mapping the type of score/remark to the EXACT column header in your Google Sheet.
# IMPORTANT: UPDATE THIS DICTIONARY TO MATCH YOUR GOOGLE SHEET HEADERS AND SEMESTER/SUBJECT STRUCTURE.
SUBJECT_DETAILS = {
    'Semester 1': {
        'Math': {
            'Exams Score': 'mathematics', # Replace 'Math Exams Score' with the actual column header
            'Class Score': 'math_class_score', # Replace 'Math Class Score' with the actual column header
            'Total Score': 'math_total', # Replace 'Math Total Score' with the actual column header
            'Grade' : 'math_grade',
            'Remarks': 'math_remark'               # Replace with actual column header
        },
        'Science': {
            'Exams Score': 'science', # Replace with actual column header
            'Class Score': 'science_class_score', # Replace with actual column header
            'Total Score': 'science_total', # Replace with actual column header
            'Grade' : 'science_grade',
            'Remarks': 'science_remark'          # Replace with actual column header
        },
        'English': {
            'Exams Score': 'english', # Replace with actual column header
            'Class Score': 'english_class_score', # Replace with actual column header
            'Total Score': 'english_total', # Replace with actual column header
            'Grade' : 'english_grade',
            'Remarks': 'english_remark'
        },
        'Social': {
            'Exams Score': 'social', # Replace with actual column header
            'Class Score': 'social_class_score', # Replace with actual column header
            'Total Score': 'social_total', # Replace with actual column header
            'Grade' : 'social_grade',
            'Remarks': 'social_remark'
        },
        'Entrepreneur': {
            'Exams Score': 'entrepreneur', # Replace with actual column header
            'Class Score': 'entrepreneur_class_score', # Replace with actual column header
            'Total Score': 'entrepreneur_total', # Replace with actual column header
            'Grade' : 'entrepreneur_grade',
            'Remarks': 'entrepreneur_remark'
        },
        'ICT': {
            'Exams Score': 'ict', # Replace with actual column header
            'Class Score': 'ict_class_score', # Replace with actual column header
            'Total Score': 'ict_total', # Replace with actual column header
            'Grade' : 'ict_grade',
            'Remarks': 'ict_remark'
        },
        'Technical Drawing': {
            'Exams Score': 'td', # Replace with actual column header
            'Class Score': 'td_class_score', # Replace with actual column header
            'Total Score': 'td_total', # Replace with actual column header
            'Grade' : 'technical_grade',
            'Remarks': 'td_remark'
        },
        'Elective Course': {
            'Exams Score': 'electivecourse', # Replace with actual column header
            'Class Score': 'electivecourse_class_score', # Replace with actual column header
            'Total Score': 'electivecourse_total', # Replace with actual column header
            'Elective Course': 'elective_course',
            'Grade': 'elective_grade',
            'Remarks': 'electivecourse_remark'
        },
        # Add more Semester 1 subjects here
    },
    
    'Semester 2': {
        'Math': {
            'Exams Score': 'mathematics', # Replace 'Math Exams Score' with the actual column header
            'Class Score': 'math_class_score', # Replace 'Math Class Score' with the actual column header
            'Total Score': 'math_total', # Replace 'Math Total Score' with the actual column header
            ' Grade': 'math_grade',
            'Remarks': 'math_remark'             # Replace with actual column header
        },
        'Science': {
            'Exams Score': 'science', # Replace with actual column header
            'Class Score': 'science_class_score', # Replace with actual column header
            'Total Score': 'science_total', # Replace with actual column header
            'Grade': 'science_grade',
            'Remarks': 'science_remark'           # Replace with actual column header
        },
        # 'English': {
        #     'Exams Score': 'english', # Replace with actual column header
        #     'Class Score': 'english_class_score', # Replace with actual column header
        #     'Total Score': 'english_total', # Replace with actual column header
        #     'Remarks': 'english_remark'
        # },
        'English': {
            'Exams Score': 'english', # Replace with actual column header
            'Class Score': 'english_class_score', # Replace with actual column header
            'Total Score': 'english_total', # Replace with actual column header
            'Grade': 'english_grade',
            'Remarks': 'english_remark'
        },
        'Social': {
            'Exams Score': 'social', # Replace with actual column header
            'Class Score': 'social_class_score', # Replace with actual column header
            'Total Score': 'social_total', # Replace with actual column header
            'Grade': 'social_grade',
            'Remarks': 'social_remark'
        },
        'Entrepreneur': {
            'Exams Score': 'entrepreneur', # Replace with actual column header
            'Class Score': 'entrepreneur_class_score', # Replace with actual column header
            'Total Score': 'entrepreneur_total', # Replace with actual column header
            'Grade': 'entrepreneur_grade',
            'Remarks': 'entrepreneur_remark'
        },
        'ICT': {
            'Exams Score': 'ict', # Replace with actual column header
            'Class Score': 'ict_class_score', # Replace with actual column header
            'Total Score': 'ict_total', # Replace with actual column header
            'Grade' : 'ict_grade',
            'Remarks': 'ict_remark'
        },
        'Technical Drawing': {
            'Exams Score': 'td', # Replace with actual column header
            'Class Score': 'td_class_score', # Replace with actual column header
            'Total Score': 'td_total', # Replace with actual column header
            'Grade' : 'td_grade',
            'Remarks': 'td_remark'
        },
        'Elective Course': {
            'Exams Score': 'electivecourse', # Replace with actual column header
            'Class Score': 'electivecourse_class_score', # Replace with actual column header
            'Total Score': 'electivecourse_total', # Replace with actual column header
            'Grade' : 'electivecourse_grade',
            'Remarks': 'electivecourse_remark'
        },
        # Add more Semester 2 subjects here
    },
    # Add more semesters here following the same structure
}

# Populate COLUMN_MAPPING with subject-specific columns based on SUBJECT_DETAILS
for semester, subjects in SUBJECT_DETAILS.items():
    for subject, details in subjects.items():
        for key, col_name in details.items():
             # Create a unique internal key (e.g., 'Semester 1 - Math Exams Score')
            COLUMN_MAPPING[f'{semester} - {subject} {key}'] = col_name


# --- Helper to get student data by ID ---
def get_student_data_by_id(student_id):
    """Loads data and finds a student by ID."""
    df = load_results_from_sheet()
    if df.empty:
        return None, "Could not load results data."

    # Ensure the Student ID column is treated as string for reliable matching
    if COLUMN_MAPPING['Student ID'] not in df.columns:
         return None, f"Required column '{COLUMN_MAPPING['Student ID']}' not found in sheet."

    df[COLUMN_MAPPING['Student ID']] = df[COLUMN_MAPPING['Student ID']].astype(str).str.strip()

    student_row = df[df[COLUMN_MAPPING['Student ID']] == student_id.strip()]

    if student_row.empty:
        return None, f"Student ID {student_id} not found."

    return student_row.iloc[0].to_dict(), None # Return student data and no error


# --- Data Loading Function ---
def load_results_from_sheet():
    """Loads the student results from the Google Sheet into a pandas DataFrame."""
    try:
        # Read the CSV data directly from the export URL
        df = pd.read_csv(CSV_EXPORT_URL)
        # Optional: Clean up column names by removing leading/trailing spaces
        df.columns = df.columns.str.strip()

        # --- Verify essential columns exist immediately after loading ---
        essential_cols = [COLUMN_MAPPING['Student ID'], COLUMN_MAPPING['Student Name'], COLUMN_MAPPING['Parent Phone']]
        if not all(col in df.columns for col in essential_cols):
            missing = [col for col in essential_cols if col not in df.columns]
            print(f"Error: Missing ESSENTIAL columns in sheet: {missing}")
            # Flash a message if possible, though this function might be called before request context is fully set up
            # flash(f"Error: Missing essential columns in sheet: {missing}", 'danger')
            return pd.DataFrame() # Return empty DataFrame if essential columns are missing
        # --- End essential column verification ---


        # Optional: Verify that all mapped columns exist (including subject details)
        all_mapped_cols = list(COLUMN_MAPPING.values())
        missing_mapped = [col for col in all_mapped_cols if col not in df.columns]
        if missing_mapped:
            print(f"Warning: Missing some MAPPED columns in sheet: {missing_mapped}")
            # Continue, but data for these columns will be 'N/A'


        # --- Fix for phone numbers ending in .0 ---
        # Ensure the Parent Phone column exists before attempting to clean it (already checked above, but defensive)
        if COLUMN_MAPPING['Parent Phone'] in df.columns:
            df[COLUMN_MAPPING['Parent Phone']] = df[COLUMN_MAPPING['Parent Phone']].astype(str).str.replace('.0', '', regex=False).str.strip()
            print("Cleaned Parent Phone column.")
        # --- End of fix ---


        print("Successfully loaded data from Google Sheet.")
        return df
    except Exception as e:
        print(f"Error loading data from Google Sheet: {e}")
        # flash(f"Error loading data from Google Sheet: {e}", 'danger') # Flash error if possible
        return pd.DataFrame() # Return empty DataFrame on error


# --- Arkesel SMS Function ---
def send_sms(phone_number, message):
    """Sends an SMS message using the Arkesel API (Older GET endpoint)."""
    # Ensure phone number is in a valid format for Arkesel (e.g., starts with country code)
    # Basic cleaning: remove spaces and dashes
    cleaned_phone = str(phone_number).replace(" ", "").replace("-", "")
    # Add country code if missing (assuming Ghana +233). Adjust if needed.
    if not cleaned_phone.startswith('+'):
         # This is a simple assumption, you might need more sophisticated logic
         if cleaned_phone.startswith('0'):
             cleaned_phone = '+233' + cleaned_phone[1:] # Replace leading 0 with +233
         else:
             cleaned_phone = '+233' + cleaned_phone # Prepend +233

    # Validate phone number format (basic check)
    if not cleaned_phone or len(cleaned_phone) < 10: # Minimum length check
         print(f"Invalid phone number format: {phone_number}")
         return False, "Invalid phone number format."

    # --- Construct payload as URL parameters for the older GET endpoint ---
    payload = {
        'action': 'send-sms',
        'api_key': ARKESEL_API_KEY,
        'to': cleaned_phone,
        'from': ARKESEL_SENDER_ID,
        'sms': message
    }
    print(f"Attempting to send SMS to {cleaned_phone} with message: {message}")
    print(f"SMS API Payload (URL Params): {payload}") # Print the payload being sent

    try:
        # Use requests.get for the older endpoint and pass params
        response = requests.get(ARKESEL_SMS_URL, params=payload)

        # --- Debugging: Print status code and raw response text ---
        print(f"SMS API HTTP Status Code: {response.status_code}")
        print(f"SMS API Raw Response Text: {response.text}")
        # --- End Debugging ---

        # The older endpoint might return plain text or a different format, not always JSON.
        # We'll check for a successful status code (200) and look for indicators of success in the text.
        # You might need to adjust this success check based on actual Arkesel response for this endpoint.
        if response.status_code == 200:
            # Assuming success is indicated by a specific string in the response text
            # Replace 'SUCCESS_INDICATOR_STRING' with the actual string Arkesel returns on success
            # Common indicators might be 'OK', 'success', a specific code, etc.
            # If the response is just the message ID, checking for a non-empty text might suffice.
            if response.text and ("OK" in response.text.upper() or response.text.isdigit()): # Example check: adjust based on real response
                 return True, "SMS sent successfully!"
            else:
                 # If status is 200 but text doesn't indicate success, use the raw text as error
                 return False, f"SMS failed: API returned 200 but response indicates failure - {response.text}"
        else:
            # Handle non-200 status codes
            return False, f"SMS failed: HTTP Status Code {response.status_code} - {response.text}"

    except requests.exceptions.RequestException as e:
        print(f"Network error sending SMS: {e}")
        return False, f"Network error sending SMS: {e}"
    except Exception as e:
        print(f"An unexpected error occurred during SMS sending: {e}")
        return False, f"An unexpected error occurred during SMS sending: {e}"


# --- Flask Routes ---

@app.route('/')
def index():
    """School website homepage with login options."""
    return render_template('index.html')

@app.route('/student_login', methods=['GET', 'POST'])
def student_login():
    """Handles student login via name and parent contact."""
    if request.method == 'POST':
        student_name = request.form.get('student_name')
        parent_phone = request.form.get('parent_phone')

        print(f"Student login attempt: Name='{student_name}', Phone='{parent_phone}'") # Debug print

        if not student_name or not parent_phone:
            flash("Please enter both student name and parent contact.", 'warning')
            return render_template('student_login.html') # Render the student login form

        df = load_results_from_sheet()
        if df.empty:
             flash("Could not load results data. Please try again later.", 'danger')
             print("Error: Dataframe is empty after load_results_from_sheet in student_login.") # Debug print
             return render_template('student_login.html') # Render the student login form

        # Find the student by matching name and phone number
        # Use case-insensitive comparison for name, strip whitespace
        # Use stripped phone number for comparison, including cleaning .0
        required_student_cols = [COLUMN_MAPPING['Student Name'], COLUMN_MAPPING['Parent Phone']]
        if not all(col in df.columns for col in required_student_cols):
             missing_cols = [col for col in required_student_cols if col not in df.columns]
             flash(f"Required columns for verification not found in sheet: {missing_cols}.", 'danger')
             print(f"Error: Missing required student login columns: {missing_cols} in student_login.") # Debug print
             return render_template('student_login.html')

        df['_temp_name'] = df[COLUMN_MAPPING['Student Name']].astype(str).str.strip().str.lower()
        df['_temp_phone'] = df[COLUMN_MAPPING['Parent Phone']].astype(str).str.strip().replace(" ", "").replace("-", "").replace(".0", "", regex=False)

        student_row = df[
            (df['_temp_name'] == student_name.strip().lower()) &
            (df['_temp_phone'] == parent_phone.strip().replace(" ", "").replace("-", "")) # Clean input phone for comparison
        ]

        # Drop temporary columns
        df = df.drop(columns=['_temp_name', '_temp_phone'])


        if student_row.empty:
            flash("Invalid student name or parent contact.", 'danger')
            print(f"Student not found for Name='{student_name}', Phone='{parent_phone}' in student_login.") # Debug print
            return render_template('student_login.html') # Render the student login form

        # Assuming the first match is the correct student
        student_data = student_row.iloc[0].to_dict()
        print(f"Student found: {student_data.get(COLUMN_MAPPING['Student Name'])}") # Debug print

        # Prepare data for display, including all subject details per semester
        display_results = {
            'Student Name': student_data.get(COLUMN_MAPPING['Student Name'], 'N/A'),
            'Student ID': student_data.get(COLUMN_MAPPING['Student ID'], 'N/A'),
            'Semesters': {} # Nested dictionary for semesters
        }

        # Populate subject details per semester for display
        for semester, subjects in SUBJECT_DETAILS.items():
            display_results['Semesters'][semester] = {} # Create nested dictionary for semester
            for subject, details in subjects.items():
                subject_info = {}
                for key, col_name in details.items():
                     # Check if the column exists before trying to get the value
                     if col_name in student_data:
                        subject_info[key] = student_data.get(col_name, 'N/A')
                     else:
                        subject_info[key] = 'Column Missing' # Indicate missing column in data
                        print(f"Warning: Column '{col_name}' for subject '{subject}' in semester '{semester}' not found in student data.") # Debug print

                display_results['Semesters'][semester][subject] = subject_info

        print(f"Prepared display_results for student: {display_results.get('Student Name')}") # Debug print


        # Redirect to the results page (or render it directly)
        # For simplicity, let's render the results page directly after successful login
        return render_template('results.html', student_data=display_results)

    # Handle GET request - display student login form
    return render_template('student_login.html')


@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    """Handles admin login via password."""
    if request.method == 'POST':
        password = request.form.get('password')

        # Check the submitted password against the stored hash
        if check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['admin_logged_in'] = True # Set session variable on successful login
            flash('Logged in successfully!', 'success')
            return redirect(url_for('admin_dashboard')) # Redirect to admin dashboard
        else:
            flash('Invalid password.', 'danger')
            return render_template('admin_login.html') # Render login form with error

    # Handle GET request - display admin login form
    return render_template('admin_login.html')

@app.route('/admin')
def admin_dashboard():
    """Admin dashboard to view all results and trigger SMS (protected)."""
    # Check if admin is logged in
    if not session.get('admin_logged_in'):
        flash('Please log in to access the admin dashboard.', 'warning')
        return redirect(url_for('admin_login'))

    df = load_results_from_sheet()
    if df.empty:
        return render_template('admin.html', error="Could not load results data for admin view. Check sheet access.")

    # Convert DataFrame to list of dictionaries for easy rendering in HTML
    # Include all mapped columns for admin view
    display_columns = list(COLUMN_MAPPING.values())
    # Ensure only existing columns are selected
    existing_display_columns = [col for col in display_columns if col in df.columns]

    results_list = df[existing_display_columns].to_dict('records')

    # Flash messages are handled by the base template

    return render_template('admin.html', results=results_list)


@app.route('/admin/logout')
def admin_logout():
    """Logs out the admin user."""
    session.pop('admin_logged_in', None) # Remove session variable
    flash('Logged out successfully.', 'success')
    return redirect(url_for('admin_login'))


@app.route('/admin/send_sms/<student_id>')
def admin_send_single_sms(student_id):
    """Sends an SMS with results for a specific student (protected)."""
    # Check if admin is logged in
    if not session.get('admin_logged_in'):
        flash('Please log in to perform this action.', 'warning')
        return redirect(url_for('admin_login'))

    print(f"Received student_id in admin_send_single_sms route: {student_id}") # Debug print

    student_data, error_msg = get_student_data_by_id(student_id)

    if error_msg:
        flash(error_msg, 'danger')
        return redirect(url_for('admin_dashboard'))

    phone_number = str(student_data.get(COLUMN_MAPPING['Parent Phone'], '')).strip()
    student_name = str(student_data.get(COLUMN_MAPPING['Student Name'], '')).strip()

    if not phone_number:
        flash(f"No phone number found for {student_name} ({student_id}). Cannot send SMS.", 'warning')
        return redirect(url_for('admin_dashboard'))

    # --- Construct the SMS message with all subject scores and remarks on separate lines per semester ---
    message_lines = [f"Dear Parent of {student_name},"] # Start with the greeting

    # Iterate through semesters defined in SUBJECT_DETAILS
    for semester, subjects in SUBJECT_DETAILS.items():
        message_lines.append(f"\n--- {semester} ---") # Add semester header

        # Iterate through subjects within this semester
        for subject, details in subjects.items():
            # Use .get() with a default to avoid KeyError if a column is missing
            exams_score = student_data.get(details.get('Exams Score'), 'N/A')
            class_score = student_data.get(details.get('Class Score'), 'N/A')
            total_score = student_data.get(details.get('Total Score'), 'N/A')
            remarks = student_data.get(details.get('Remarks'), 'N/A')

            # Format the line for this subject
            subject_line = f"{subject}: Exams:{exams_score}, Class:{class_score}, Total:{total_score}, Remarks:{remarks}"
            message_lines.append(subject_line) # Add each subject line

    message_body = "\n".join(message_lines) # Join lines with newline


    # Construct the URL for the student results page (PDF version)
    # IMPORTANT: URL encode the parameters, especially name, to handle spaces and special characters
    encoded_name = urllib.parse.quote_plus(student_name)
    encoded_phone = urllib.parse.quote_plus(phone_number) # Phone number should ideally be cleaned/formatted consistently

    # Construct the full URL using the configured WEBSITE_DOMAIN
    results_link = f"{WEBSITE_DOMAIN}{url_for('student_result_pdf', name=encoded_name, phone=encoded_phone)}"
    print(f"Generated results_link URL: {results_link}") # Debug print for URL generation


    # Combine message body and link
    full_message = f"{message_body}\n\nView PDF results: {results_link}" # Added extra newline for separation
    print(f"Generated SMS: {full_message}") # Print the full message to console


    success, msg = send_sms(phone_number, full_message)
    flash(msg, 'success' if success else 'danger') # Flash the SMS result message

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/send_all_sms')
def admin_send_all_sms():
    """Sends SMS results to all parents listed in the sheet (protected)."""
    # Check if admin is logged in
    if not session.get('admin_logged_in'):
        flash('Please log in to perform this action.', 'warning')
        return redirect(url_for('admin_login'))

    df = load_results_from_sheet()
    if df.empty:
        flash("Error loading data to send SMS to all.", 'danger')
        return redirect(url_for('admin_dashboard'))

    sent_count = 0
    failed_count = 0
    messages = [] # Collect messages to flash at the end

    # Iterate through each student row
    for index, student_data in df.iterrows():
        student_id = str(student_data.get(COLUMN_MAPPING['Student ID'], '')).strip()
        student_name = str(student_data.get(COLUMN_MAPPING['Student Name'], '')).strip()
        phone_number = str(student_data.get(COLUMN_MAPPING['Parent Phone'], '')).strip()


        if not phone_number:
            messages.append(f"Skipping SMS for {student_name} ({student_id}): No phone number found.")
            failed_count += 1
            continue

        # --- Construct the SMS message with all subject scores and remarks on separate lines per semester ---
        message_lines = [f"Dear Parent of {student_name},"] # Start with the greeting

        # Iterate through semesters defined in SUBJECT_DETAILS
        for semester, subjects in SUBJECT_DETAILS.items():
            message_lines.append(f"\n--- {semester} ---") # Add semester header

            # Iterate through subjects within this semester
            for subject, details in subjects.items():
                # Use .get() with a default to avoid KeyError if a column is missing
                exams_score = student_data.get(details.get('Exams Score'), 'N/A')
                class_score = student_data.get(details.get('Class Score'), 'N/A')
                total_score = student_data.get(details.get('Total Score'), 'N/A')
                remarks = student_data.get(details.get('Remarks'), 'N/A')

                # Format the line for this subject
                subject_line = f"{subject}: Exams:{exams_score}, Class:{class_score}, Total:{total_score}, Remarks:{remarks}"
                message_lines.append(subject_line) # Add each subject line

        message_body = "\n".join(message_lines) # Join lines with newline


        # Construct the URL for the student results page (PDF version)
        # IMPORTANT: URL encode the parameters
        encoded_name = urllib.parse.quote_plus(student_name)
        encoded_phone = urllib.parse.quote_plus(phone_number)

        # Construct the full URL using the configured WEBSITE_DOMAIN
        results_link = f"{WEBSITE_DOMAIN}{url_for('student_result_pdf', name=encoded_name, phone=encoded_phone)}"
        print(f"Generated results_link URL for {student_name}: {results_link}") # Debug print for URL generation

        # Combine message body and link
        full_message = f"{message_body}\n\nView PDF results: {results_link}" # Added extra newline for separation
        print(f"Generated SMS for {student_name}: {full_message}") # Print the full message to console


        success, msg = send_sms(phone_number, full_message)
        if success:
            messages.append(f"SMS sent to {student_name}: {msg}")
            sent_count += 1
        else:
            messages.append(f"SMS failed for {student_name}: {msg}")
            failed_count += 1

    status_message = f"SMS sending batch complete. Sent: {sent_count}, Failed: {failed_count}. See messages below for details."
    flash(status_message, 'info') # Flash the summary message
    # Flash individual messages - consider if this is too much for many students
    # for msg in messages:
    #     flash(msg, 'info')


    return redirect(url_for('admin_dashboard'))


@app.route('/student_result_pdf')
def student_result_pdf():
    """Generates a PDF report of a student's result using name and phone for verification."""
    student_name = request.args.get('name')
    parent_phone = request.args.get('phone')

    if not student_name or not parent_phone:
        return "Missing student name or phone number.", 400 # Bad Request

    df = load_results_from_sheet()
    if df.empty:
         return "Could not load results data.", 500 # Internal Server Error

    # Find the student by matching name and phone number
    if COLUMN_MAPPING['Student Name'] not in df.columns or COLUMN_MAPPING['Parent Phone'] not in df.columns:
         return "Required columns for verification not found in sheet.", 500

    df['_temp_name'] = df[COLUMN_MAPPING['Student Name']].astype(str).str.strip().str.lower()
    # Ensure the phone number from the sheet is cleaned consistently with the input
    df['_temp_phone'] = df[COLUMN_MAPPING['Parent Phone']].astype(str).str.strip().replace(" ", "").replace("-", "").replace(".0", "", regex=False)

    student_row = df[
        (df['_temp_name'] == student_name.strip().lower()) &
        (df['_temp_phone'] == parent_phone.strip().replace(" ", "").replace("-", "")) # Clean input phone for comparison
    ]

    df = df.drop(columns=['_temp_name', '_temp_phone'])

    if student_row.empty:
        return "Could not retrieve results. Please check the link or contact the school.", 404 # Not Found

    student_data = student_row.iloc[0].to_dict()

    # Prepare data for display in PDF template, including all subject details per semester
    display_results = {
        'Student Name': student_data.get(COLUMN_MAPPING['Student Name'], 'N/A'),
        'Student ID': student_data.get(COLUMN_MAPPING['Student ID'], 'N/A'),
        'Semesters': {} # Nested dictionary for semesters
    }

    # Populate subject details per semester for display
    for semester, subjects in SUBJECT_DETAILS.items():
        display_results['Semesters'][semester] = {} # Create nested dictionary for semester
        for subject, details in subjects.items():
            subject_info = {}
            for key, col_name in details.items():
                 # Check if the column exists before trying to get the value
                 if col_name in student_data:
                    subject_info[key] = student_data.get(col_name, 'N/A')
                 else:
                    subject_info[key] = 'Column Missing' # Indicate missing column in data
                    print(f"Warning: Column '{col_name}' for subject '{subject}' in semester '{semester}' not found in student data for PDF.") # Debug print

            display_results['Semesters'][semester][subject] = subject_info


    # Render the HTML template designed for PDF
    # Pass the prepared data and current time to the template
    rendered_html = render_template('student_result_pdf.html', student_data=display_results, now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    # Generate PDF from the rendered HTML
    pdf = HTML(string=rendered_html).write_pdf()

    # Create a Flask response with the PDF
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    # Suggest a filename for the downloaded PDF
    filename = f"{display_results['Student Name'].replace(' ', '_')}_Results.pdf"
    response.headers['Content-Disposition'] = f'inline; filename={filename}' # Use 'inline' to display in browser, 'attachment' to download

    return response

@app.route('/ai_assistant', methods=['GET', 'POST'])
def ai_assistant():
    """AI Assistant page for student research."""
    search_results = None
    answer = None
    query = request.form.get('query') if request.method == 'POST' else None
    document_content_summary = None # Summary or list of documents

    if request.method == 'POST' and query:
        print(f"AI Assistant received query: {query}")

        # --- Simulate using document content (requires parsing libraries) ---
        # In a real scenario, you would process the query and potentially search
        # within the uploaded documents if the query is relevant.
        # For this simulation, we'll just list available documents or indicate their presence.
        available_documents = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if os.path.isfile(os.path.join(app.config['UPLOAD_FOLDER'], f)) and allowed_file(f)]
        if available_documents:
            document_content_summary = "Considering information from the school library (available documents: " + ", ".join(available_documents) + ")."
            # In a real AI, you would send relevant document chunks to the LLM based on the query.
        else:
             document_content_summary = "No documents available in the school library."
        # --- End simulation of document content usage ---


        # --- Perform web search using the tool ---
        if search: # Check if the search tool is available
            try:
                # The search tool returns a list of SearchResults objects
                search_results_list = search(queries=[query])
                if search_results_list and search_results_list[0].results:
                    search_results = search_results_list[0].results
                    print(f"Search results: {search_results}")

                    # --- Simulate LLM processing of query, search results, and document content ---
                    # This is a basic simulation. A real LLM would process the query,
                    # search results, and relevant document content to generate an answer.
                    answer_lines = [f"AI Assistant response to: '{query}'"]
                    if document_content_summary:
                        answer_lines.append(f"\n({document_content_summary})")

                    if search_results:
                        answer_lines.append("\nBased on web search results:")
                        for i, res in enumerate(search_results[:3]): # Limit to top 3 results for brevity
                            answer_lines.append(f"{i+1}. {res.snippet or 'No snippet available'} ({res.source_title or 'Unknown Source'})")
                    else:
                        answer_lines.append("Could not find relevant information from web search.")

                    # Combine simulated answer parts
                    answer = "\n".join(answer_lines)
                    # --- End simulation ---

                else:
                    answer = "Could not find relevant information from web search."
                    print("No search results returned.")

            except Exception as e:
                print(f"Error during AI Assistant search: {e}")
                answer = f"An error occurred during search: {e}"
        else:
            answer = "Web search tool is not available."
            if document_content_summary:
                 answer += f" {document_content_summary}"


    return render_template('ai_assistant.html', search_results=search_results, answer=answer, query=query)

@app.route('/library', methods=['GET', 'POST'])
def library():
    """Library page for admin uploads and student access (admin protected for upload)."""
    files = []
    try:
        # List only allowed files in the uploads directory
        files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if os.path.isfile(os.path.join(app.config['UPLOAD_FOLDER'], f)) and allowed_file(f)]
    except Exception as e:
        print(f"Error listing files in library: {e}")
        flash(f"Error loading library files: {e}", 'danger')

    # Check if admin is logged in for upload functionality
    is_admin = session.get('admin_logged_in', False)

    if request.method == 'POST':
        # Only process upload if admin is logged in
        if not is_admin:
            flash('You do not have permission to upload files.', 'danger')
            return redirect(url_for('library'))

        # Handle file upload
        if 'file' not in request.files:
            flash('No file part in the request.', 'warning')
            return redirect(url_for('library'))

        file = request.files['file']

        if file.filename == '':
            flash('No selected file.', 'warning')
            return redirect(url_for('library'))

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            try:
                file.save(filepath)
                flash(f'File "{filename}" uploaded successfully.', 'success')
            except Exception as e:
                print(f"Error saving file {filename}: {e}")
                flash(f'Error uploading file "{filename}": {e}', 'danger')
        else:
            flash('Invalid file type. Allowed types are: pdf, doc, docx, xls, xlsx.', 'warning')

        # Redirect back to the library page after upload
        return redirect(url_for('library'))

    # Handle GET request
    return render_template('library.html', files=files, is_admin=is_admin)

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Serve uploaded files."""
    try:
        # Ensure the requested file is within the allowed extensions before serving
        if not allowed_file(filename):
             return "File type not allowed.", 403 # Forbidden

        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except FileNotFoundError:
        return "File not found.", 404


if __name__ == '__main__':
    # In a production environment, use a production-ready WSGI server like Gunicorn or uWSGI
    # For local development, this is fine
    # Set debug=False for production
    app.run(debug=True)
